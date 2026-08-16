#!/usr/bin/env python3
"""Build jobs.xlsx from the built bilingual dataset (data/jobs.zh.json).

Run after build_dataset.py, which produces data/jobs.zh.json from the raw
snapshot and the DeepSeek translations.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def _cell(value: Any) -> Any:
    """Keep cells visually empty when the source field is missing."""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return int(value)
    return value


def _yes_no(value: bool) -> str:
    return "是" if value else "否"


_STATUS_EMOJI = {
    "是": "✅",
    "否": "❌",
    "必需": "✅",
    "可选/非必需": "🟡",
    "未明确要求": "❌",
    "明确要求/需核验": "✅",
    "入职后可取得": "🟡",
    "未发现明确要求": "❌",
    "明确要求经验年限": "✅",
    "未发现明确年限": "❌",
}


def _status_emoji(status: str) -> str:
    return _STATUS_EMOJI.get(status, status)


_EMOJI_COLOR = {"✅": "FF1E7B1E", "🟡": "FFBF8F00", "❌": "FFC00000"}


def _color_emoji_cells(ws: Worksheet, cols: Sequence[int]) -> None:
    for row_idx in range(5, ws.max_row + 1):
        for col in cols:
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell.value, str):
                color = _EMOJI_COLOR.get(cell.value)
                if color:
                    cell.font = Font(color=color)


def _style_title(ws: Worksheet, ncols: int, title: str, subtitle: str) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1, value=subtitle)
    ws.cell(row=2, column=1).font = Font(color="808080", size=10)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].height = 6


def _style_header(ws: Worksheet, headers: Sequence[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="BFBFBF")
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin)
    ws.freeze_panes = "A5"


def _set_widths(ws: Worksheet, widths: Sequence[tuple[int, int]]) -> None:
    for col, width in widths:
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_rows(ws: Worksheet, rows: Sequence[Sequence[Any]], wrap_cols: set[int] = frozenset()) -> None:
    thin = Side(style="thin", color="D9D9D9")
    for row in rows:
        ws.append([_cell(value) for value in row])
    for row_idx in range(5, 5 + len(rows)):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col)
            if col in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=thin)


def _link_title(ws: Worksheet, jobs: Sequence[dict[str, Any]], title_col: int) -> None:
    """Turn each title cell into a hyperlink to the official posting."""
    for i, job in enumerate(jobs):
        cell = ws.cell(row=5 + i, column=title_col)
        cell.hyperlink = job["url"]
        cell.font = Font(color="0563C1", underline="single")


def build_search_sheet(wb: Workbook, jobs: Sequence[dict[str, Any]], fetched_at: str) -> None:
    ws = wb.active
    ws.title = "岗位检索"
    headers = [
        "序号", "中文部门", "中文职位", "开放日期", "截止日期",
        "每周标准工时", "最低周工时", "最高周工时", "薪资原文", "最低时薪", "最高时薪",
        "Work-Study", "本科生限定", "驾照", "食品处理员许可证", "经验要求",
        "额外要求", "页面明确关闭", "班次", "Requisition Number", "GUID",
        "英文部门", "英文职位",
    ]
    _style_title(ws, len(headers), "犹他大学 2026 Fall 校内兼职检索表",
                 f"共 {len(jobs)} 个兼职岗位 · 快照 {fetched_at} · 全量保留，不按个人身份或具体申请日期删岗")
    _style_header(ws, headers)
    rows = []
    for index, r in enumerate(jobs, 1):
        rows.append([
            index, r["department_zh"], r["title_zh"],
            r["open_date"], r["close_date"],
            r["standard_hours_text"], r["minimum_weekly_hours"], r["maximum_weekly_hours"],
            r["pay_rate_text"], r["minimum_hourly_pay"], r["maximum_hourly_pay"],
            _status_emoji(r["work_study_status"]),
            _status_emoji(_yes_no(r["undergraduate_only"])),
            _status_emoji(r["driver_license_status"]),
            _status_emoji(r["food_handler_status"]),
            _status_emoji(r["experience_status"]),
            r["extra_requirements"],
            _status_emoji(_yes_no(r["explicitly_closed_in_posting_text"])),
            r["shift"],
            r["requisition_number"], r["guid"],
            r["department_en"], r["title_en"],
        ])
    _write_rows(ws, rows, wrap_cols={2, 3, 17, 22, 23})
    _link_title(ws, jobs, title_col=3)
    _color_emoji_cells(ws, cols=[12, 13, 14, 15, 16, 18])
    ws.auto_filter.ref = f"A4:{ws.cell(row=ws.max_row, column=len(headers)).coordinate}"
    _set_widths(ws, [
        (1, 6), (2, 22), (3, 34), (4, 11), (5, 11),
        (6, 14), (7, 10), (8, 10), (9, 12), (10, 9), (11, 9),
        (12, 11), (13, 10), (14, 8), (15, 14), (16, 10),
        (17, 30), (18, 11), (19, 10), (20, 13), (21, 34),
        (22, 30), (23, 30),
    ])


def build_chinese_sheet(wb: Workbook, jobs: Sequence[dict[str, Any]]) -> None:
    ws = wb.create_sheet("完整中文")
    headers = [
        "序号", "中文部门", "中文职位", "职位摘要", "工作职责", "最低资格", "优先条件", "申请说明",
        "排班安排", "工时", "薪资", "开放日期", "截止日期", "GUID",
    ]
    _style_title(ws, len(headers), "完整中文内容",
                 "每个岗位按官网字段统一拆分；没有内容的字段保留为空。中文用于检索和速读，申请前请核对英文原文。")
    _style_header(ws, headers)
    rows = []
    for index, r in enumerate(jobs, 1):
        rows.append([
            index, r["department_zh"], r["title_zh"],
            r["job_summary_zh"], r["responsibilities_zh"], r["minimum_qualifications_zh"],
            r["preferences_zh"], r["special_instructions_zh"], r["work_schedule_summary_zh"],
            r["standard_hours_text"], r["pay_rate_text"], r["open_date"], r["close_date"], r["guid"],
        ])
    _write_rows(ws, rows, wrap_cols={2, 3, 4, 5, 6, 7, 8, 9})
    _link_title(ws, jobs, title_col=3)
    _set_widths(ws, [
        (1, 6), (2, 22), (3, 34), (4, 60), (5, 60), (6, 60), (7, 60),
        (8, 50), (9, 50), (10, 14), (11, 12), (12, 11), (13, 11), (14, 34),
    ])


def build_english_sheet(wb: Workbook, jobs: Sequence[dict[str, Any]]) -> None:
    ws = wb.create_sheet("英文原文")
    headers = [
        "No.", "Department", "Job Title", "Job Summary", "Responsibilities",
        "Minimum Qualifications", "Preferences", "Special Instructions", "Work Schedule",
        "Hours", "Pay", "Open Date", "Close Date", "GUID",
    ]
    _style_title(ws, len(headers), "Complete English Source",
                 "Original text extracted from the saved Jobsyn snapshot. Columns align with the Complete Chinese sheet.")
    _style_header(ws, headers)
    rows = []
    for index, r in enumerate(jobs, 1):
        rows.append([
            index, r["department_en"], r["title_en"],
            r["job_summary_en"], r["responsibilities_en"], r["minimum_qualifications_en"],
            r["preferences_en"], r["special_instructions_en"], r["work_schedule_summary_en"],
            r["standard_hours_text"], r["pay_rate_text"], r["open_date"], r["close_date"], r["guid"],
        ])
    _write_rows(ws, rows, wrap_cols={2, 3, 4, 5, 6, 7, 8, 9})
    _link_title(ws, jobs, title_col=3)
    _set_widths(ws, [
        (1, 6), (2, 30), (3, 34), (4, 60), (5, 60), (6, 60), (7, 60),
        (8, 50), (9, 50), (10, 14), (11, 12), (12, 11), (13, 11), (14, 34),
    ])


def build_manual_sheet(wb: Workbook, jobs: Sequence[dict[str, Any]]) -> None:
    ws = wb.create_sheet("使用说明")
    ncols = 6
    _style_title(ws, ncols, "使用说明",
                 f"公开版保留全部 {len(jobs)} 个兼职岗位。标签只帮助缩小范围，不替申请人或学校判断资格。")

    work_study_required = sum(1 for r in jobs if r["work_study_status"] == "必需")
    undergrad = sum(1 for r in jobs if r["undergraduate_only"])
    closed = sum(1 for r in jobs if r["explicitly_closed_in_posting_text"])

    row = 4
    ws.cell(row=row, column=1, value="数据概览").font = Font(bold=True)
    ws.cell(row=row, column=2, value="数值").font = Font(bold=True)
    row += 1
    overview = [
        ("兼职岗位总数", len(jobs)),
        ("Work-Study 必需", work_study_required),
        ("明确本科生限定", undergrad),
        ("页面明确关闭", closed),
        ("完整中文字段", "摘要、职责、最低资格、优先条件、申请说明、排班"),
    ]
    for label, value in overview:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row=row, column=1, value="状态图例").font = Font(bold=True)
    row += 1
    legend = [
        ("✅", "是 / 必需 / 明确要求"),
        ("🟡", "可选 / 入职后可取得"),
        ("❌", "否 / 未发现（未发现不代表确认无要求）"),
    ]
    for mark, meaning in legend:
        ws.cell(row=row, column=1, value=mark)
        ws.cell(row=row, column=2, value=meaning)
        row += 1

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row=row, column=1, value="建议用法").font = Font(bold=True)
    row += 1
    usage = [
        ("1", "先看岗位检索", "按自己的申请日期、身份、工时、薪资和证照条件筛选，不使用预设个人档案。"),
        ("2", "再读完整中文", "逐栏查看摘要、职责、最低资格、优先条件、申请说明和排班。"),
        ("3", "核对英文原文", "中文翻译用于快速阅读；涉及资格、证照、日期和申请材料时回英文列核对。"),
        ("4", "打开职位链接", "中文职位名即为官网链接，直接点击；快照不是实时页面，正式申请前确认仍开放。"),
        ("5", "自行排序", "可按截止日期、最低时薪、最低周工时、部门或任何资格标签重新排序。"),
    ]
    for num, label, desc in usage:
        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=desc)
        row += 1

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row=row, column=1, value="标签口径").font = Font(bold=True)
    row += 1
    notes = [
        ("❌ 未发现明确要求", "仅表示自动规则没有识别到，不是学校确认没有。"),
        ("❌ 本科生限定 = 否", "表示没有发现明确限定，不代表研究生一定符合。"),
        ("✅ 页面明确关闭", "正文写明不再接受申请，即使截止日期在未来也先跳过。"),
        ("教育折抵经验", "仍需结合专业和职位给出的折抵公式判断。"),
        ("翻译", "完整中文由 AI 辅助生成；申请前以英文原文为准。"),
    ]
    for label, desc in notes:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=desc)
        row += 1

    _set_widths(ws, [(1, 16), (2, 40), (3, 60), (4, 20), (5, 20), (6, 20)])


def _intl_flags(r: dict[str, Any]) -> str:
    flags: list[str] = []
    if r["driver_license_status"] == "明确要求/需核验":
        flags.append("需驾照")
    elif r["driver_license_status"] == "入职后可取得":
        flags.append("驾照可后考")
    if r["food_handler_status"] == "明确要求/需核验":
        flags.append("需食品证")
    elif r["food_handler_status"] == "入职后可取得":
        flags.append("食品证可后考")
    if r["experience_status"] == "明确要求经验年限":
        flags.append("需经验")
    if r["extra_requirements"]:
        flags.append(r["extra_requirements"])
    return "；".join(flags)


def build_intl_sheet(wb: Workbook, jobs: Sequence[dict[str, Any]], fetched_at: str) -> None:
    """MEAE 国际学生省流版：排除必须联邦勤工助学（FWS）、仅限本科、已关闭、需公民身份的岗位。"""
    ws = wb.create_sheet("MEAE国际学生省流版")
    headers = ["序号", "职位", "部门", "每周工时", "时薪", "截止日期", "注意"]
    eligible = [
        r for r in jobs
        if r["work_study_status"] != "必需"
        and not r["undergraduate_only"]
        and not r["explicitly_closed_in_posting_text"]
        and not r["requires_citizenship"]
    ]
    excluded = len(jobs) - len(eligible)
    _style_title(ws, len(headers), "MEAE 国际学生省流版",
                 f"面向 MEAE 硕士国际学生（F-1）· 已排除 {excluded} 个必须勤工助学、仅限本科、已关闭或需公民身份的岗位 · 快照 {fetched_at}")
    _style_header(ws, headers)
    # 截止日期近的在前，未提供截止日期的排最后
    ordered = sorted(eligible, key=lambda r: (r["close_date"] is None, r["close_date"] or ""))
    rows = []
    for index, r in enumerate(ordered, 1):
        pay = r["minimum_hourly_pay"]
        pay_text = ("$%g" % pay) if pay is not None else (r["pay_rate_text"] or "")
        rows.append([
            index, r["title_zh"], r["department_zh"],
            r["standard_hours_text"], pay_text, r["close_date"],
            _intl_flags(r),
        ])
    _write_rows(ws, rows, wrap_cols={2, 3, 7})
    _link_title(ws, ordered, title_col=2)
    ws.auto_filter.ref = f"A4:{ws.cell(row=ws.max_row, column=len(headers)).coordinate}"
    _set_widths(ws, [(1, 6), (2, 30), (3, 22), (4, 14), (5, 10), (6, 11), (7, 40)])


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    root = Path(__file__).resolve().parents[1]
    parser.add_argument("--json-input", type=Path, default=root / "data/jobs.zh.json")
    parser.add_argument("--output", type=Path, default=root / "jobs.xlsx")
    args = parser.parse_args()

    dataset = json.loads(args.json_input.read_text(encoding="utf-8"))
    jobs: list[dict[str, Any]] = dataset["jobs"]
    fetched_at: str = dataset["snapshot_fetched_at"]

    wb = Workbook()
    build_search_sheet(wb, jobs, fetched_at)
    build_chinese_sheet(wb, jobs)
    build_english_sheet(wb, jobs)
    build_intl_sheet(wb, jobs, fetched_at)
    build_manual_sheet(wb, jobs)

    assert len(jobs) == dataset["job_count"]
    wb.save(args.output)
    print(f"Built {len(jobs)} jobs -> {args.output.name} ({', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    main()
