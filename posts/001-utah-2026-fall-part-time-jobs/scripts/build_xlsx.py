#!/usr/bin/env python3
"""Build jobs.xlsx from the built bilingual dataset (data/jobs.zh.json).

Run after build_dataset.py, which produces data/jobs.zh.json from the raw
snapshot and the DeepSeek translations.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


def _cell(value: Any) -> Any:
    """Keep cells visually empty when the source field is missing."""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return int(value)
    return value


def _yes_no(value: bool) -> str:
    return "是" if value else "否"


def _style_title(ws: Worksheet, ncols: int, title: str, subtitle: str) -> None:
    """Write the merged title row and subtitle row, then a blank spacer row."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1, value=subtitle)
    ws.cell(row=2, column=1).font = Font(color="808080", size=10)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # row 3 stays blank; header goes on row 4.
    ws.row_dimensions[3].height = 6


def _style_header(ws: Worksheet, headers: Sequence[str]) -> int:
    header_row = 4
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="BFBFBF")
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin)
    ws.freeze_panes = "A5"
    return header_row


def _set_widths(ws: Worksheet, widths: Sequence[tuple[int, int]]) -> None:
    from openpyxl.utils import get_column_letter

    for col, width in widths:
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_rows(ws: Worksheet, rows: Sequence[Sequence[Any]], wrap_cols: set[int] = frozenset()) -> None:
    thin = Side(style="thin", color="D9D9D9")
    for row in rows:
        ws.append([_cell(value) for value in row])
    for row_idx in range(5, 5 + len(rows)):
        for col in wrap_cols:
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(1, ws.max_column + 1):
            current = ws.cell(row=row_idx, column=col)
            if col in wrap_cols:
                current.alignment = Alignment(wrap_text=True, vertical="top")
            current.border = Border(bottom=thin)


def build_search_sheet(wb: Workbook, jobs: Sequence[dict[str, Any]], fetched_at: str) -> None:
    ws = wb.active
    ws.title = "岗位检索"
    headers = [
        "序号", "中文部门", "英文部门", "中文职位", "英文职位",
        "开放日期", "截止日期", "优先审查日", "每周标准工时", "最低周工时", "最高周工时",
        "薪资原文", "最低时薪", "最高时薪", "Work-Study", "本科生限定",
        "驾照", "食品处理员许可证", "酒类服务证书", "经验要求", "F-1工时风险",
        "页面明确关闭", "班次", "官网链接", "Requisition Number", "GUID",
    ]
    _style_title(ws, len(headers), "犹他大学 2026 Fall 校内兼职检索表",
                 f"共 {len(jobs)} 个兼职岗位 · 快照 {fetched_at} · 全量保留，不按个人身份或具体申请日期删岗")
    _style_header(ws, headers)
    rows = []
    for index, r in enumerate(jobs, 1):
        rows.append([
            index, r["department_zh"], r["department_en"], r["title_zh"], r["title_en"],
            r["open_date"], r["close_date"], r["priority_review_date"],
            r["standard_hours_text"], r["minimum_weekly_hours"], r["maximum_weekly_hours"],
            r["pay_rate_text"], r["minimum_hourly_pay"], r["maximum_hourly_pay"],
            r["work_study_status"], _yes_no(r["undergraduate_only"]),
            r["driver_license_status"], r["food_handler_status"], r["alcohol_certificate_status"],
            r["experience_status"], _yes_no(r["f1_hours_risk"]),
            _yes_no(r["explicitly_closed_in_posting_text"]),
            r["shift"], r["url"], r["requisition_number"], r["guid"],
        ])
    _write_rows(ws, rows, wrap_cols={2, 3, 4, 5, 24})
    ws.auto_filter.ref = f"A4:{ws.cell(row=ws.max_row, column=len(headers)).coordinate}"
    _set_widths(ws, [
        (1, 6), (2, 22), (3, 30), (4, 30), (5, 30),
        (6, 11), (7, 11), (8, 11), (9, 14), (10, 10), (11, 10),
        (12, 12), (13, 9), (14, 9), (15, 12), (16, 10),
        (17, 14), (18, 16), (19, 14), (20, 14), (21, 10),
        (22, 11), (23, 10), (24, 46), (25, 13), (26, 34),
    ])


def build_chinese_sheet(wb: Workbook, jobs: Sequence[dict[str, Any]]) -> None:
    ws = wb.create_sheet("完整中文")
    headers = [
        "序号", "中文部门", "中文职位", "职位摘要", "工作职责", "最低资格", "优先条件", "申请说明",
        "排班安排", "补充信息", "工时", "薪资", "开放日期", "截止日期", "官网链接", "GUID",
    ]
    _style_title(ws, len(headers), "完整中文内容",
                 "每个岗位按官网字段统一拆分；没有内容的字段保留为空。中文用于检索和速读，申请前请核对英文原文。")
    _style_header(ws, headers)
    rows = []
    for index, r in enumerate(jobs, 1):
        rows.append([
            index, r["department_zh"], r["title_zh"],
            r["job_summary_zh"], r["responsibilities_zh"], r["minimum_qualifications_zh"],
            r["preferences_zh"], r["special_instructions_zh"], r["work_schedule_summary_zh"],
            r["additional_information_zh"], r["standard_hours_text"], r["pay_rate_text"],
            r["open_date"], r["close_date"], r["url"], r["guid"],
        ])
    _write_rows(ws, rows, wrap_cols={2, 3, 4, 5, 6, 7, 8, 9, 10, 15})
    _set_widths(ws, [
        (1, 6), (2, 22), (3, 30), (4, 60), (5, 60), (6, 60), (7, 60), (8, 50),
        (9, 50), (10, 50), (11, 14), (12, 12), (13, 11), (14, 11), (15, 46), (16, 34),
    ])


def build_english_sheet(wb: Workbook, jobs: Sequence[dict[str, Any]]) -> None:
    ws = wb.create_sheet("英文原文")
    headers = [
        "No.", "Department", "Job Title", "Job Summary", "Responsibilities",
        "Minimum Qualifications", "Preferences", "Special Instructions", "Work Schedule",
        "Additional Information", "Hours", "Pay", "Open Date", "Close Date", "Job URL", "GUID",
    ]
    _style_title(ws, len(headers), "Complete English Source",
                 "Original text extracted from the saved Jobsyn snapshot. Columns align with the Complete Chinese sheet.")
    _style_header(ws, headers)
    rows = []
    for index, r in enumerate(jobs, 1):
        rows.append([
            index, r["department_en"], r["title_en"],
            r["job_summary_en"], r["responsibilities_en"], r["minimum_qualifications_en"],
            r["preferences_en"], r["special_instructions_en"], r["work_schedule_summary_en"],
            r["additional_information_en"], r["standard_hours_text"], r["pay_rate_text"],
            r["open_date"], r["close_date"], r["url"], r["guid"],
        ])
    _write_rows(ws, rows, wrap_cols={2, 3, 4, 5, 6, 7, 8, 9, 10, 15})
    _set_widths(ws, [
        (1, 6), (2, 30), (3, 30), (4, 60), (5, 60), (6, 60), (7, 60), (8, 50),
        (9, 50), (10, 50), (11, 14), (12, 12), (13, 11), (14, 11), (15, 46), (16, 34),
    ])


def build_manual_sheet(wb: Workbook, jobs: Sequence[dict[str, Any]]) -> None:
    ws = wb.create_sheet("使用说明")
    ncols = 6
    _style_title(ws, ncols, "使用说明",
                 f"公开版保留全部 {len(jobs)} 个兼职岗位。标签只帮助缩小范围，不替申请人或学校判断资格。")

    work_study_required = sum(1 for r in jobs if r["work_study_status"] == "必需")
    undergrad = sum(1 for r in jobs if r["undergraduate_only"])
    closed = sum(1 for r in jobs if r["explicitly_closed_in_posting_text"])
    f1_risk = sum(1 for r in jobs if (r["minimum_weekly_hours"] or 0) > 20)

    ws.cell(row=4, column=1, value="数据概览").font = Font(bold=True)
    ws.cell(row=4, column=2, value="数值").font = Font(bold=True)
    overview = [
        ("兼职岗位总数", len(jobs)),
        ("Work-Study 必需", work_study_required),
        ("明确本科生限定", undergrad),
        ("页面明确关闭", closed),
        ("最低周工时 > 20", f1_risk),
        ("完整中文字段", "摘要、职责、最低资格、优先条件、申请说明、排班、补充信息"),
    ]
    for i, (label, value) in enumerate(overview):
        ws.cell(row=5 + i, column=1, value=label)
        ws.cell(row=5 + i, column=2, value=value)

    ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=ncols)
    ws.cell(row=13, column=1, value="建议用法").font = Font(bold=True)
    usage = [
        ("1", "先看岗位检索", "按自己的申请日期、身份、工时、薪资和证照条件筛选，不使用预设个人档案。"),
        ("2", "再读完整中文", "逐栏查看摘要、职责、最低资格、优先条件、申请说明、排班和补充信息。"),
        ("3", "核对英文原文", "中文翻译用于快速阅读；涉及资格、证照、日期和申请材料时回英文列核对。"),
        ("4", "打开官网链接", "快照不是实时页面。正式申请前确认职位仍开放，且官网内容没有更新。"),
        ("5", "自行排序", "可按截止日期、最低时薪、最低周工时、部门或任何资格标签重新排序。"),
    ]
    for i, (num, label, desc) in enumerate(usage):
        ws.cell(row=14 + i, column=1, value=num)
        ws.cell(row=14 + i, column=2, value=label)
        ws.cell(row=14 + i, column=3, value=desc)

    ws.merge_cells(start_row=21, start_column=1, end_row=21, end_column=ncols)
    ws.cell(row=21, column=1, value="标签口径").font = Font(bold=True)
    notes = [
        ("未发现明确要求", "仅表示自动规则没有识别到，不是学校确认没有。"),
        ("本科生限定 = 否", "表示没有发现明确限定，不代表研究生一定符合。"),
        ("教育折抵经验", "仍需结合专业和职位给出的折抵公式判断。"),
        ("F-1 工时风险", "仅以最低周工时是否超过 20 作保守提示，不构成移民或法律意见。"),
        ("翻译", "完整中文由 AI 辅助生成；申请前以英文原文为准。"),
    ]
    for i, (label, desc) in enumerate(notes):
        ws.cell(row=22 + i, column=1, value=label)
        ws.cell(row=22 + i, column=2, value=desc)

    _set_widths(ws, [(1, 16), (2, 18), (3, 60), (4, 20), (5, 20), (6, 20)])


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    root = Path(__file__).resolve().parents[1]
    parser.add_argument("--json-input", type=Path, default=root / "data/jobs.zh.json")
    parser.add_argument("--output", type=Path, default=root / "jobs.xlsx")
    args = parser.parse_args()

    dataset = json.loads(args.json_input.read_text(encoding="utf-8"))
    jobs: list[dict[str, Any]] = dataset["jobs"]
    fetched_at: str = dataset["snapshot_fetched_at"]

    wb = Workbook()
    build_search_sheet(wb, jobs, fetched_at)
    build_chinese_sheet(wb, jobs)
    build_english_sheet(wb, jobs)
    build_manual_sheet(wb, jobs)

    assert len(jobs) == dataset["job_count"]
    wb.save(args.output)
    print(f"Built {len(jobs)} jobs -> {args.output.name} ({', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    main()
