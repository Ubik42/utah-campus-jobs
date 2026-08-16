import importlib.util
import json
import unittest
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = Path(__file__).with_name("build_xlsx.py")
SPEC = importlib.util.spec_from_file_location("build_xlsx", SCRIPT)
build_xlsx = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
SPEC.loader.exec_module(build_xlsx)


class XlsxTests(unittest.TestCase):
    def setUp(self) -> None:
        dataset = json.loads((ROOT / "data/jobs.zh.json").read_text(encoding="utf-8"))
        self.jobs = dataset["jobs"]
        self.count = dataset["job_count"]

    def build(self):
        wb = Workbook()
        build_xlsx.build_search_sheet(wb, self.jobs, "snapshot")
        build_xlsx.build_chinese_sheet(wb, self.jobs)
        build_xlsx.build_english_sheet(wb, self.jobs)
        build_xlsx.build_manual_sheet(wb, self.jobs)
        return wb

    def test_four_sheets_and_data_rows(self) -> None:
        wb = self.build()
        self.assertEqual(wb.sheetnames, ["岗位检索", "完整中文", "英文原文", "使用说明"])
        self.assertEqual(wb["岗位检索"].max_row, 4 + self.count)
        self.assertEqual(wb["完整中文"].max_row, 4 + self.count)
        self.assertEqual(wb["英文原文"].max_row, 4 + self.count)

    def test_search_sheet_headers_and_filter(self) -> None:
        wb = self.build()
        ws = wb["岗位检索"]
        headers = [cell.value for cell in ws[4]]
        self.assertEqual(headers[0], "序号")
        self.assertEqual(headers[2], "中文职位")
        self.assertEqual(headers[-2:], ["英文部门", "英文职位"])
        self.assertIn("GUID", headers)
        self.assertIsNotNone(ws.auto_filter.ref)

    def test_status_emoji_mapping(self) -> None:
        self.assertEqual(build_xlsx._status_emoji("必需"), "✅")
        self.assertEqual(build_xlsx._status_emoji("入职后可取得"), "🟡")
        self.assertEqual(build_xlsx._status_emoji("未发现明确要求"), "❌")

    def test_manual_sheet_stats_match_data(self) -> None:
        wb = self.build()
        ws = wb["使用说明"]
        self.assertEqual(ws.cell(row=5, column=2).value, self.count)
        required = sum(1 for r in self.jobs if r["work_study_status"] == "必需")
        self.assertEqual(ws.cell(row=6, column=2).value, required)


if __name__ == "__main__":
    unittest.main()
